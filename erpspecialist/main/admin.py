from django.contrib import admin
from django.urls import path
from django.shortcuts import render
from django import forms
from openpyxl import load_workbook
from .models import ByYearStatistic, ByCityStatistic, SkillStatistic, Images


class DataUploadForm(forms.Form):
    data_upload = forms.FileField()


class ByYearStatAdmin(admin.ModelAdmin):
    def get_urls(self):
        url = super().get_urls()
        new_url = [path('data-upload/', self.data_upload, name='data_upload')]
        return new_url + url

    def data_upload(self, req):
        if req.method == 'POST':
            wb = load_workbook(filename=req.FILES['data_upload'].file)
            for row in wb.active:
                values = [cell.value for cell in row]
                created = ByYearStatistic.objects.update_or_create(
                    year=values[0],
                    salary_by_year=values[1],
                    count_by_year=values[2],
                    vac_salary_by_year=values[3],
                    vac_count_by_year=values[4]
                )

        form = DataUploadForm()
        data = {'form': form}
        return render(req, 'admin/data_upload.html', data)


class ByCityStatAdmin(admin.ModelAdmin):
    def get_urls(self):
        url = super().get_urls()
        new_url = [path('data-upload/', self.data_upload, name='data_upload')]
        return new_url + url

    def data_upload(self, req):
        if req.method == 'POST':
            wb = load_workbook(filename=req.FILES['data_upload'].file)
            for row in wb.active:
                values = [cell.value for cell in row]
                created = ByCityStatistic.objects.update_or_create(
                    city_salary=values[0],
                    salary=values[1],
                    city_part=values[2],
                    part=values[3]
                )

        form = DataUploadForm()
        data = {'form': form}
        return render(req, 'admin/data_upload.html', data)


class SkillStatAdmin(admin.ModelAdmin):
    def get_urls(self):
        url = super().get_urls()
        new_url = [path('data-upload/', self.data_upload, name='data_upload')]
        return new_url + url

    def data_upload(self, req):
        if req.method == 'POST':
            wb = load_workbook(filename=req.FILES['data_upload'].file)
            for row in wb.active:
                values = [cell.value for cell in row]
                created = SkillStatistic.objects.update_or_create(
                    year=values[0],
                    skills=values[1]
                )

        form = DataUploadForm()
        data = {'form': form}
        return render(req, 'admin/data_upload.html', data)


admin.site.register(ByYearStatistic, ByYearStatAdmin)
admin.site.register(ByCityStatistic, ByCityStatAdmin)
admin.site.register(SkillStatistic, SkillStatAdmin)
admin.site.register(Images)
