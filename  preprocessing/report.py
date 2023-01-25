import matplotlib.pyplot as plt
from utils import Utils
from typing import Dict, List
from openpyxl import Workbook


class Report:
    def __init__(self):
        self.workbook_year = Workbook()
        self.workbook_city = Workbook()
        self.workbook_skills = Workbook()

    def generate_excel(
            self,
            all_salary_level: Dict[str, int],
            all_vacancies_count: Dict[str, int],
            salary_level: Dict[str, int],
            vacancies_count: Dict[str, int],
            by_city_level: Dict[str, float],
            vacancies_part: Dict[str, float],
            year_skills: Dict[str, List[str]]
    ) -> None:
        self.__make_year(
            self.workbook_year.active,
            all_salary_level,
            all_vacancies_count,
            salary_level,
            vacancies_count
        )
        self.__make_city(
            self.workbook_city.active,
            Utils.slice_dict(by_city_level, 10),
            Utils.slice_dict(vacancies_part, 10)
        )
        self.__make_skills(
            self.workbook_skills.active,
            year_skills
        )

        self.workbook_year.save("output/year.xlsx")
        self.workbook_city.save("output/city.xlsx")
        self.workbook_skills.save("output/skills.xlsx")

    def generate_image(
            self,
            vacancy_name: str,
            all_salary_level: Dict[str, int],
            all_vacancies_count: Dict[str, int],
            salary_level: Dict[str, int],
            vacancies_count: Dict[str, int],
            by_city_level: Dict[str, float],
            vacancies_part: Dict[str, float],
            year_skills: Dict[str, List[str]]
    ) -> None:
        fig = plt.figure(dpi=400, figsize=[17, 12])
        salary_year = fig.add_subplot(331)
        vacancies_year = fig.add_subplot(332)
        salary_city = fig.add_subplot(333)
        part_city = fig.add_subplot(334)
        skills = fig.add_subplot(335)

        salary_year.bar(
            [int(e) - 0.2 for e in all_salary_level.keys()],
            list(all_salary_level.values()),
            width=0.4,
            label="средняя з/п"
        )
        salary_year.bar(
            [int(e) + 0.2 for e in salary_level.keys()],
            list(salary_level.values()),
            width=0.4,
            label=f"з/п {vacancy_name}"
        )

        vacancies_year.bar(
            [int(e) - 0.2 for e in all_vacancies_count.keys()],
            list(all_vacancies_count.values()),
            width=0.4,
            label="Количество вакансий"
        )
        vacancies_year.bar(
            [int(e) + 0.2 for e in vacancies_count.keys()],
            list(vacancies_count.values()),
            width=0.4,
            label=f"Количество вакансий {vacancy_name}"
        )

        by_city_level_sliced = dict(reversed(Utils.slice_dict(by_city_level, 10).items()))
        salary_city.barh(
            list(by_city_level_sliced.keys()),
            list(by_city_level_sliced.values())
        )

        vacancies_part_sliced = Utils.slice_dict(vacancies_part, 10)
        other_part = sum(Utils.dict_difference(vacancies_part_sliced, vacancies_part).values())
        part_city.pie(
            [e * 100 for e in vacancies_part_sliced.values()] + [other_part * 100],
            labels=list(vacancies_part_sliced.keys()) + ["Другие"],
            textprops={'fontsize': 6}
        )

        x = [key for key, value in year_skills.items() for i in value]
        y = [item for sublist in year_skills.values() for item in sublist]
        skills.scatter(x, y, s=20, c='#1f77b4', marker='s')

        self.__format_histograms(
            salary_year,
            vacancies_year,
            salary_city,
            part_city,
            skills
        )

        fig.tight_layout(pad=7.1)

        extent1 = self.__get_bound(salary_year, fig)
        extent2 = self.__get_bound(vacancies_year, fig)
        extent3 = self.__get_bound(salary_city, fig)
        extent4 = self.__get_bound(part_city, fig)
        extent5 = self.__get_bound(skills, fig)

        fig.savefig('output/year1.jpeg', bbox_inches=extent1)
        fig.savefig('output/year2.jpeg', bbox_inches=extent2)
        fig.savefig('output/city1.jpeg', bbox_inches=extent3)
        fig.savefig('output/city2.jpeg', bbox_inches=extent4)
        fig.savefig('output/skills.jpeg', bbox_inches=extent5)

    @staticmethod
    def __get_bound(plot, fig):
        return plot.get_window_extent().transformed(fig.dpi_scale_trans.inverted()).expanded(1.4, 1.4)

    @staticmethod
    def __format_histograms(
            salary_year,
            vacancies_year,
            salary_city,
            part_city,
            skills
    ) -> None:
        salary_year.set_title("Уровень зарплат по годам")
        salary_year.grid(axis='y')
        salary_year.tick_params(labelsize=8)
        salary_year.set_xticks(salary_year.get_xticks(), salary_year.get_xticklabels(), rotation=90)
        salary_year.legend(fontsize=8)

        vacancies_year.set_title("Количество вакансий по годам")
        vacancies_year.grid(axis='y')
        vacancies_year.tick_params(labelsize=8)
        vacancies_year.set_xticks(vacancies_year.get_xticks(), vacancies_year.get_xticklabels(), rotation=90)
        vacancies_year.legend(fontsize=8)

        salary_city.grid(axis='x')
        salary_city.set_yticklabels(
            labels=[Utils.wrap_text(e.get_text(), [" ", "-"]) for e in salary_city.get_yticklabels()],
            fontsize=6
        )
        salary_city.tick_params(axis="x", labelsize=8)
        salary_city.set_title("Уровень зарплат по городам")

        part_city.set_title("Доля вакансий по городам")

        skills.set_title('ТОП-10 навыков по годам')
        skills.tick_params(labelsize=4)
        skills.grid(True, axis='y')

    @staticmethod
    def __make_year(
            worksheet,
            all_salary_level: Dict[str, int],
            all_vacancies_count: Dict[str, int],
            salary_level: Dict[str, int],
            vacancies_count: Dict[str, int]
    ) -> None:
        for key in all_salary_level.keys():
            worksheet.append([
                key,
                all_salary_level[key],
                all_vacancies_count[key],
                salary_level[key],
                vacancies_count[key]
            ])

    @staticmethod
    def __make_city(
            worksheet,
            by_city_level: Dict[str, float],
            vacancies_part: Dict[str, float]
    ) -> None:
        level_keys = list(by_city_level.keys())
        part_keys = list(vacancies_part.keys())
        for i in range(len(level_keys)):
            level_key = level_keys[i]
            part_key = part_keys[i]
            worksheet.append([
                level_key,
                by_city_level[level_key],
                part_key,
                vacancies_part[part_key]
            ])

    @staticmethod
    def __make_skills(worksheet, year_skills: Dict[str, List[str]]):
        for key in year_skills.keys():
            worksheet.append([
                key,
                ', '.join(year_skills[key])
            ])
